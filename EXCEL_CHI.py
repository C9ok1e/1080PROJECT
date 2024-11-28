import os
import time
from openpyxl import load_workbook
import pandas as pd

#PLEASE RUN #5 AND #6 IN TERMINAL IF PROGRAM WAS FAILED TO RUN
#pip install pandas openpyxl
#pip install openpyxl

data = {
    "NAME": [],
    "SUBJECT": [],
    "LANGUAGE": [],
    "CONTENT": []
}

print("歡迎使用學生資料匹配系統：資料輸入部分，請輸入你的個人資料")
print(f"當前時間: {time.strftime("%H:%M,%b%d")}")
while True:
    CHECK = input("按下任意按鈕開始程序，在最後輸入‘Done’上傳：")
    if CHECK.lower() == "done":
        break

    name= input("輸入你的名稱：").strip()
    subject= input("輸入你的科目：").strip()
    language= input("輸入你的語言：").strip()

    while True:
        content = input("輸入你的電話號碼用作聯繫（8位數字）：").strip()
        if content == "":
            content = None
            break
        elif content.isdigit() and len(content) == 8:
            break
        else:
            print("你的聯繫方式需是8位數字電話號碼，請留空或再次輸入：")
    if name or subject or language or content:
        data["NAME"].append(name.upper())
        data["SUBJECT"].append(subject.upper())
        data["LANGUAGE"].append(language.upper())
        data["CONTENT"].append(content if content else None)

if not any(data[key] for key in data):
    print("未有數據輸入，項目將不會被儲存。")
else:
    new_df = pd.DataFrame(data)
    filename = "STUDENT_INFO.xlsx"

    if os.path.exists(filename):
        existing_df = pd.read_excel(filename)
        combined_df = pd.concat([existing_df, new_df], ignore_index=True)
        combined_df.to_excel(filename, index=False)
        print("你的資料已成功更新到現有數據庫。")
    else:
        new_df.to_excel(filename, index=False)
        print("你的資料已成功儲存到系統。")

print(" ")
print("你要現在尋找你的學習夥伴嗎？")
a = input("如有需要，請輸入’yes‘進入匹配系統")

import subprocess
print(a)

if a == "yes":

    part2 = "正在連結到匹配系統"
    for i in range(4):  # Repeat 4 times to show the animation
        print(f"\r{part2}{'.' * i}", end='', flush=True)
        time.sleep(0.5)  # Pause for half a second
    print()  # Move to the next line

    try:
        subprocess.run(["python", "MATCHING_CHI.py"], check=True)
    except subprocess.CalledProcessError as e:
        print("Error executing the script: {e}")
else:
    print("似乎你暫時不需要，程序終止。")
    print("多謝你填寫資料到數據庫")
    file_path = "STUDENT_INFO.xlsx"
    sheet_name = "Sheet1"
    def count_rows_in_sheet(path, name):
        workbook = load_workbook(path)
        worksheet = workbook[name]
        return worksheet.max_row
    print(f"現時有 {count_rows_in_sheet(file_path, sheet_name)} 名學生的資訊")
