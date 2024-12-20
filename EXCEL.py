import os
import time
from openpyxl import load_workbook
import pandas as pd

#PLEASE RUN #5 AND #6 IN TERMINAL IF PROGRAM WAS FAILED TO RUN
#pip install pandas openpyxl
#pip install openpyxl

data = {
    "NAME": [],
    "SUBJECT": [],
    "LANGUAGE": [],
    "CONTENT": []
}

print("Welcome to the student matching system ! Please enter the student's information")
print(f"Current time: {time.strftime("%H:%M,%b%d")}")
while True:
    CHECK = input("PRESS ANY BUTTON TO upload your profile and type 'Done' in submit：")
    if CHECK.lower() == "done":
        break

    name= input("INPUT NAME：").strip()
    subject= input("INPUT SUBJECT：").strip()
    language= input("INPUT LANGUAGE：").strip()

    while True:
        content = input("INPUT CONTENT (MUST BE 8 DIGITS)：").strip()
        if content == "":
            content = None
            break
        elif content.isdigit() and len(content) == 8:
            break
        else:
            print("CONTENT must be exactly 8 digits or left empty. Please try again.")
    if name or subject or language or content:
        data["NAME"].append(name.upper())
        data["SUBJECT"].append(subject.upper())
        data["LANGUAGE"].append(language.upper())
        data["CONTENT"].append(content if content else None)

if not any(data[key] for key in data):
    print("No data entered, no data saved.")
else:
    new_df = pd.DataFrame(data)
    filename = "STUDENT_INFO.xlsx"

    if os.path.exists(filename):
        existing_df = pd.read_excel(filename)
        combined_df = pd.concat([existing_df, new_df], ignore_index=True)
        combined_df.to_excel(filename, index=False)
        print("INFO was successfully added to the existing file.")
    else:
        new_df.to_excel(filename, index=False)
        print("Your information was successfully saved to the system.")

print(" ")
print("Do you want to find groupmates with the system?")
a = input("If yes, please input 'yes' to match find your group mates:")

import subprocess
print(a)

if a == "yes":

    part2 = "Now going to the matching system"
    for i in range(4):  # Repeat 4 times to show the animation
        print(f"\r{part2}{'.' * i}", end='', flush=True)
        time.sleep(0.5)  # Pause for half a second
    print()  # Move to the next line

    try:
        subprocess.run(["python", "MATCHING.py"], check=True)
    except subprocess.CalledProcessError as e:
        print("Error executing the script: {e}")
else:
    print("Since you didn't input yes, The progress end")
    print("Thank you for your information")
    file_path = "STUDENT_INFO.xlsx"
    sheet_name = "Sheet1"
    def count_rows_in_sheet(path, name):
        workbook = load_workbook(path)
        worksheet = workbook[name]
        return worksheet.max_row
    print(f"Now we have {count_rows_in_sheet(file_path, sheet_name)} Student's data in total")



